"""
exportar.py — Exportación de datos a Excel
============================================
Rutas para exportar pedidos individuales, listos e historial.
"""
from flask import Blueprint, send_file, request
from database import get_db
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from services.tiempo import calcular_tiempo_laboral

exportar_bp = Blueprint('exportar', __name__)


def _crear_excel_lista(pedidos, titulo="Registros"):
    """Genera un archivo Excel con la lista de pedidos proporcionada."""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo
    ws.sheet_view.showGridLines = False

    headers = [
        "MARCA", "BULTOS", "TIPO", "RETIRADO POR", "FECHA", "ESTADO",
        "INICIO EMPAQUE", "FIN EMPAQUE", "RETIRADO EN",
        "T. REGULAR", "T. EXTRA", "T. TOTAL"
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="0D2A6E")
    band_fill   = PatternFill(fill_type="solid", fgColor="EEF1FA")
    white_fill  = PatternFill(fill_type="solid", fgColor="FFFFFF")
    thin        = Side(style="thin", color="C8D0E5")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Calibri", size=10, bold=True, color="F5C800")
    data_font   = Font(name="Calibri", size=10, color="1A1F36")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 34
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for p in pedidos:
        tipo_label = "Empaque" if p["tipo"] == "empaque" else "Directo"
        tiempo = calcular_tiempo_laboral(p.get("inicio_empaque"), p.get("fin_empaque"))
        ws.append([
            p["marca"], p.get("bultos"), tipo_label, p.get("retirado_por") or "",
            p["fecha"], p["estado"], p.get("inicio_empaque") or "",
            p.get("fin_empaque") or "", p.get("retirado_en") or "",
            tiempo["texto_regulares"] if tiempo else "",
            tiempo["texto_extras"]    if tiempo else "",
            tiempo["texto_total"]     if tiempo else "",
        ])

    for row_idx in range(2, ws.max_row + 1):
        base_fill = band_fill if row_idx % 2 == 0 else white_fill
        ws.row_dimensions[row_idx].height = 22
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = center
            cell.border = border
            cell.fill = base_fill

    anchos = [24, 10, 12, 26, 16, 13, 18, 18, 18, 14, 14, 14]
    for col_letter, width in zip("ABCDEFGHIJKL", anchos):
        ws.column_dimensions[col_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@exportar_bp.route('/exportar/listos')
def exportar_listos():
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM pedidos WHERE estado='empacado' AND eliminado=0 ORDER BY fecha DESC"
    ).fetchall()
    pedidos = [dict(r) for r in rows]
    output = _crear_excel_lista(pedidos, "Listos")
    return send_file(output, download_name="bodega_listos.xlsx",
                     as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@exportar_bp.route('/exportar/retirados')
def exportar_retirados():
    conn = get_db()
    q = request.args.get("q", "").strip()
    fecha_ini = request.args.get("fecha_ini", "").strip()
    fecha_fin = request.args.get("fecha_fin", "").strip()
    tipo_filter = request.args.get("tipo", "").strip()
    
    query = "SELECT * FROM pedidos WHERE estado='retirado' AND eliminado=0"
    params = []
    if q:
        query += " AND (marca LIKE ? OR retirado_por LIKE ?)"
        params += [f"%{q}%", f"%{q}%"]
    if fecha_ini:
        query += " AND fecha >= ?"
        params.append(fecha_ini)
    if fecha_fin:
        query += " AND fecha <= ?"
        params.append(fecha_fin + " 23:59")
    if tipo_filter in ['empaque', 'directo']:
        query += " AND tipo = ?"
        params.append(tipo_filter)
        
    query += " ORDER BY retirado_en DESC"
    rows = conn.execute(query, params).fetchall()
    pedidos = [dict(r) for r in rows]
    output = _crear_excel_lista(pedidos, "Historial")
    return send_file(output, download_name="bodega_historial.xlsx",
                     as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _crear_excel_detalle(pedido, empresa):
    """Genera un reporte Excel con 100% de fidelidad al formato original, incluyendo alturas de fila y bordes correctos."""
    from datetime import datetime
    wb = Workbook()
    ws = wb.active
    ws.title = f"Pedido #{pedido.get('correlativo_mes') or pedido['id']}"
    ws.sheet_view.showGridLines = False

    # Estilos Base
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9") # Gris neutral
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    
    title_font = Font(name="Calibri", size=18, bold=True)
    sub_font   = Font(name="Calibri", size=15, bold=True)
    header_font = Font(name="Calibri", size=13, bold=True)
    label_font = Font(name="Calibri", size=11, bold=True)
    val_font   = Font(name="Calibri", size=11)
    bulto_font = Font(name="Calibri", size=16, bold=True)
    note_font  = Font(name="Calibri", size=9)
    # TODO EL EXCEL CENTRADO (Horizontal y Vertical)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Helper para aplicar bordes a rangos combinados
    def style_range(ws, range_str, border=None, fill=None, font=None, alignment=center):
        from openpyxl.utils import get_column_interval
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        
        top_left_cell = range_str.split(':')[0]
        if font: ws[top_left_cell].font = font
        if fill: ws[top_left_cell].fill = fill
        if alignment: ws[top_left_cell].alignment = alignment
        
        # Aplicar borde a cada celda del rango
        if border:
            start, end = range_str.split(':')
            start_coord = coordinate_from_string(start)
            end_coord = coordinate_from_string(end)
            start_col = column_index_from_string(start_coord[0])
            start_row = start_coord[1]
            end_col = column_index_from_string(end_coord[0])
            end_row = end_coord[1]
            
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    ws.cell(row=row, column=col).border = border

    # Configuración de Columnas (Métricas de la foto)
    for col, width in zip("ABCDEFG", [32.0, 14.0, 14.0, 14.0, 14.0, 14.0, 14.0]):
        ws.column_dimensions[col].width = width

    # Helper para formatear fechas
    def fmt_dt_short(dt_str):
        if not dt_str: return "---"
        try:
            try: dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
            except: dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
            return dt.strftime('%I:%M %p')
        except: return dt_str

    # 1. ENCABEZADO
    ws.row_dimensions[1].height = 34.0
    ws.merge_cells("A1:G1")
    ws["A1"] = empresa.get("nombre", "MEGA ENSAMBLES INT, S.A.").upper()
    style_range(ws, "A1:G1", border=border, font=title_font, alignment=center)

    ws.row_dimensions[2].height = 24.0
    ws.merge_cells("A2:G2")
    ws["A2"] = f"PEDIDO POR TRASPASO (#{pedido.get('correlativo_mes') or pedido['id']})"
    style_range(ws, "A2:G2", border=border, font=sub_font, alignment=center)

    ws.row_dimensions[3].height = 6.0

    # 2. SECCIÓN PEDIDO
    ws.row_dimensions[4].height = 26.0
    ws.merge_cells("A4:G4")
    ws["A4"] = "PEDIDO"
    style_range(ws, "A4:G4", border=border, fill=header_fill, font=header_font, alignment=center)

    ws.row_dimensions[5].height = 22.0
    ws["A5"] = "MARCA:"
    ws["A5"].font = label_font
    ws["A5"].alignment = center
    ws["A5"].border = border
    ws.merge_cells("B5:G5")
    ws["B5"] = pedido["marca"]
    style_range(ws, "B5:G5", border=border, font=val_font, alignment=Alignment(horizontal="left", vertical="center", indent=1))

    ws.row_dimensions[6].height = 24.0
    ws["A6"] = "RECIBIDO:"
    ws["A6"].font = label_font
    ws["A6"].alignment = center
    ws["A6"].border = border
    for col, val in zip("BCD", ["MES", "DIA", "AÑO"]):
        ws[col+"6"] = val
        ws[col+"6"].font = label_font
        ws[col+"6"].alignment = center
        ws[col+"6"].border = border
    ws.merge_cells("E6:G6")
    style_range(ws, "E6:G6", border=border, font=label_font, alignment=center)
    ws["E6"] = "HORA"

    ws.row_dimensions[7].height = 22.0
    ws["A7"].border = border
    ws["B7"] = str(pedido.get('recibido_mes') or "").zfill(2)
    ws["C7"] = str(pedido.get('recibido_dia') or "").zfill(2)
    ws["D7"] = str(pedido.get('recibido_ano') or "")
    ws.merge_cells("E7:G7")
    rec_hora = pedido.get('recibido_hora') or ""
    if rec_hora and ":" in rec_hora:
        try:
            h_dt = datetime.strptime(rec_hora, "%H:%M")
            rec_hora = h_dt.strftime("%I:%M %p")
        except: pass
    ws["E7"] = rec_hora
    for col in "BCD":
        ws[col+"7"].font = val_font
        ws[col+"7"].alignment = center
        ws[col+"7"].border = border
    style_range(ws, "E7:G7", border=border, font=val_font, alignment=center)

    ws.row_dimensions[8].height = 22.0
    ws["A8"] = "HOJAS:"
    ws["A8"].font = label_font
    ws["A8"].alignment = center
    ws["A8"].border = border
    ws.merge_cells("B8:G8")
    ws["B8"] = pedido.get("hojas") or ""
    style_range(ws, "B8:G8", border=border, font=val_font, alignment=center)

    # 3. SECCIONES / PASILLOS (CONSOLIDADO FICHA STYLE O GLOBAL)
    ws.row_dimensions[9].height = 24.0
    ws.merge_cells("A9:G9")
    ws["A9"] = "PREPARACIÓN GLOBAL" if pedido.get("preparador_nombre") else "SECCIONES / PASILLOS"
    style_range(ws, "A9:G9", border=border, fill=header_fill, font=header_font, alignment=center)

    if pedido.get("preparador_nombre"):
        # MODO GLOBAL: Un solo bloque grande
        ws.row_dimensions[10].height = 0 # Ocultar cabecera de pasillos
        ws.row_dimensions[11].height = 110.0
        for r in [12, 13, 14]: ws.row_dimensions[r].height = 0
        
        ws.merge_cells("A11:A14")
        ws["A11"] = "DETALLE DE\nTRABAJO GLOBAL:"
        style_range(ws, "A11:A14", border=border, font=label_font, alignment=center)
        
        ws.merge_cells("B11:G14")
        nombre = pedido["preparador_nombre"].upper()
        ini = fmt_dt_short(pedido.get("inicio_preparacion"))
        fin = fmt_dt_short(pedido.get("fin_preparacion"))
        t = pedido.get("tiempo_preparacion")
        dur = t["texto_regulares"] if (t and t["minutos_totales"] > 0) else "---"
        
        ws["B11"] = f"RESPONSABLE: {nombre}\nHORARIO: {ini} - {fin}\nDURACIÓN TOTAL: {dur}"
        style_range(ws, "B11:G14", border=border, font=val_font, alignment=center)

    elif pedido.get("modo_preparacion") == "HOJAS":
        # MODO POR HOJAS: Lista vertical
        ws["A9"] = "PREPARACIÓN POR HOJAS"
        ws.row_dimensions[10].height = 0
        for r in [12, 13, 14]: ws.row_dimensions[r].height = 0
        
        ws.merge_cells("A11:A14")
        ws["A11"] = "DETALLE DE\nTRABAJO POR HOJAS:"
        style_range(ws, "A11:A14", border=border, font=label_font, alignment=center)
        
        ws.merge_cells("B11:G14")
        
        hoja_records = pedido.get("hoja_records", {})
        fichas_list = []
        max_h = pedido.get("max_hoja_calc") or 1
        for n in range(1, max_h + 1):
             parts = sorted(hoja_records.get(n, []), key=lambda x: x.get("inicio") or "")
             for p in parts:
                 if p.get("persona") or p.get("inicio") or p.get("fin"):
                     nombre_p = (p.get("persona") or "SIN ASIGNAR").upper()
                     ini_p = fmt_dt_short(p.get("inicio"))
                     fin_p = fmt_dt_short(p.get("fin"))
                     t_p = p.get("tiempo_laboral")
                     dur_p = t_p["texto_regulares"] if (t_p and t_p["minutos_totales"] > 0) else ""
                     dur_str = f" ({dur_p})" if dur_p else ""
                     ficha = f"[Hoja {n}] {nombre_p}: {ini_p} a {fin_p}{dur_str}"
                     fichas_list.append(ficha)
        
        if not fichas_list:
            fichas_list.append("Sin registros de hojas")
            
        full_text = "\n".join(fichas_list)
        ws["B11"] = full_text
        style_range(ws, "B11:G14", border=border, font=val_font, alignment=center)
        
        lines_count = full_text.count("\n") + 1
        ws.row_dimensions[11].height = max(110.0, lines_count * 18.0)

    else:
        # MODO POR PASILLOS: Grid original
        ws.row_dimensions[10].height = 26.0
        ws["A10"].border = border
        for col, val in zip("BCDE", ["Pasillo 1", "Pasillo 2", "Pasillo 3", "Pasillo 4"]):
            ws[col+"10"] = val
            ws[col+"10"].font = label_font
            ws[col+"10"].alignment = center
            ws[col+"10"].border = border
        ws.merge_cells("F10:G10")
        ws["F10"] = "2do Piso"
        style_range(ws, "F10:G10", border=border, font=label_font, alignment=center)

        # Lógica de Ficha Consolidada para Pasillos
        secciones = pedido.get("secciones", {})
        pasillo_fichas = {}
        max_lines_global = 1

        for n in range(1, 6):
            parts = sorted(secciones.get(n, []), key=lambda x: x.get("inicio") or "")
            if not parts:
                pasillo_fichas[n] = "---"
                continue
            
            fichas_list = []
            for p in parts:
                nombre_p = (p.get("persona") or "SIN ASIGNAR").upper()
                ini_p = fmt_dt_short(p.get("inicio"))
                fin_p = fmt_dt_short(p.get("fin"))
                t_p = p.get("tiempo_laboral")
                dur_p = t_p["texto_regulares"] if (t_p and t_p["minutos_totales"] > 0) else "---"
                
                ficha = f"{nombre_p}\n{ini_p} - {fin_p}\nDuración: {dur_p}"
                fichas_list.append(ficha)
            
            full_text = "\n--------------------------\n".join(fichas_list)
            pasillo_fichas[n] = full_text
            
            lines_count = full_text.count("\n") + 1
            if lines_count > max_lines_global:
                max_lines_global = lines_count

        block_height = max(110.0, max_lines_global * 18.0)
        ws.row_dimensions[11].height = block_height
        for r in [12, 13, 14]: ws.row_dimensions[r].height = 0

        ws.merge_cells("A11:A14")
        ws["A11"] = "DETALLE DE TRABAJO\nPOR PASILLO:"
        style_range(ws, "A11:A14", border=border, font=label_font, alignment=center)

        for n in range(1, 6):
            col_start = n + 1 if n < 5 else 6
            col_end = n + 1 if n < 5 else 7
            range_str = f"{ws.cell(11, col_start).column_letter}11:{ws.cell(14, col_end).column_letter}14"
            ws.merge_cells(range_str)
            ws.cell(11, col_start).value = pasillo_fichas[n]
            style_range(ws, range_str, border=border, font=val_font, alignment=center)

    # 4. DIGITADO / DICTADO
    ws.row_dimensions[15].height = 22.0
    ws["A15"] = "DIGITADO POR:"
    ws["A15"].font = label_font
    ws["A15"].alignment = center
    ws["A15"].border = border
    ws.merge_cells("B15:D15")
    ws["B15"] = pedido.get("digitado_por") or ""
    style_range(ws, "B15:D15", border=border, font=val_font, alignment=center)
    
    ws["E15"] = "DICTADO POR:"
    ws["E15"].font = label_font
    ws["E15"].alignment = center
    ws["E15"].border = border
    ws.merge_cells("F15:G15")
    ws["F15"] = pedido.get("dictado_por") or ""
    style_range(ws, "F15:G15", border=border, font=val_font, alignment=center)

    # 5. DIGITACIÓN GRID
    def add_dig_row(row, label, key_dt, h_main, h_val):
        ws.row_dimensions[row].height = h_main
        ws.cell(row, 1, label).font = label_font
        ws.cell(row, 1).alignment = center
        ws.cell(row, 1).border = border
        for col, val in zip("BCD", ["MES", "DIA", "AÑO"]):
            ws.cell(row, ord(col)-64, val).font = label_font
            ws.cell(row, ord(col)-64).alignment = center
            ws.cell(row, ord(col)-64).border = border
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        style_range(ws, f"E{row}:G{row}", border=border, font=label_font, alignment=center)
        ws.cell(row, 5, "HORA")
        
        row += 1
        ws.row_dimensions[row].height = h_val
        d = pedido.get(key_dt)
        ws.cell(row, 1).border = border
        ws.cell(row, 2, d[5:7] if d else "").font = val_font
        ws.cell(row, 2).alignment = center
        ws.cell(row, 2).border = border
        ws.cell(row, 3, d[8:10] if d else "").font = val_font
        ws.cell(row, 3).alignment = center
        ws.cell(row, 3).border = border
        ws.cell(row, 4, d[:4] if d else "").font = val_font
        ws.cell(row, 4).alignment = center
        ws.cell(row, 4).border = border
        
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        d_hora = ""
        if d and len(d) >= 16:
            try:
                h_dt = datetime.strptime(d[11:16], "%H:%M")
                d_hora = h_dt.strftime("%I:%M %p")
            except: d_hora = d[11:16]
        ws.cell(row, 5, d_hora)
        style_range(ws, f"E{row}:G{row}", border=border, font=val_font, alignment=center)
        return row + 1

    next_r = add_dig_row(16, "DIGITACION COMIENZO:", "inicio_empaque", 24.0, 22.0)
    next_r = add_dig_row(next_r, "DIGITACION TERMINA:", "fin_empaque", 24.0, 22.0)

    # 6. TOTALES
    ws.row_dimensions[20].height = 22.0
    ws["A20"] = "DURACIÓN DE\nDIGITACIÓN:"
    ws["A20"].font = label_font
    ws["A20"].alignment = center
    ws["A20"].border = border
    ws.merge_cells("B20:G20")
    t_dig = calcular_tiempo_laboral(pedido.get("inicio_empaque"), pedido.get("fin_empaque"))
    ws["B20"] = t_dig["texto_regulares"] if (t_dig and t_dig["minutos_totales"] > 0) else "---"
    style_range(ws, "B20:G20", border=border, font=val_font, alignment=center)

    ws.row_dimensions[21].height = 36.0
    ws["A21"] = "OBSERVACIÓN:"
    ws["A21"].font = label_font
    ws["A21"].alignment = center
    ws["A21"].border = border
    ws.merge_cells("B21:G21")
    ws["B21"] = pedido.get("observacion") or ""
    style_range(ws, "B21:G21", border=border, font=val_font, alignment=center)

    ws.row_dimensions[22].height = 32.0
    ws["A22"] = "TOTAL BULTOS:"
    ws["A22"].font = label_font
    ws["A22"].alignment = center
    ws["A22"].border = border
    ws.merge_cells("B22:G22")
    ws["B22"] = pedido.get("bultos") or 0
    style_range(ws, "B22:G22", border=border, font=bulto_font, alignment=center)

    ws.row_dimensions[23].height = 8.0 # Separador
    ws.row_dimensions[24].height = 28.0
    ws.merge_cells("A24:D24")
    ws["A24"] = "TOTAL CANTIDAD BULTOS FINAL"
    style_range(ws, "A24:D24", border=border, font=label_font, alignment=center)
    
    ws.merge_cells("E24:G24")
    ws["E24"] = pedido.get("total_bultos_final") or pedido.get("bultos") or 0
    style_range(ws, "E24:G24", border=border, font=bulto_font, alignment=center)

    ws.row_dimensions[25].height = 14.0 # Separador
    curr_row = 26
    ws.row_dimensions[curr_row].height = 75.0 # Altura justa para la nota en papel
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
    ws.cell(curr_row, 1).value = (
        "NOTA SOBRE CÁLCULOS DE TIEMPO:\n"
        "El Tiempo documentado descuenta sistemáticamente el horario de almuerzo y las todas las horas transcurridas fuera del horario laboral estándar "
        "(Lunes a Viernes de 7:00 a.m. a 4:30 p.m., Vi. hasta 3:30 p.m.).\n"
        "Si percibe discordancias fuertes entre 'Inicio' y 'Fin' contra la 'Duración', es porque las noches y "
        "los fines de semana libres NO son computados como tiempo de trabajo."
    )
    style_range(ws, f"A{curr_row}:G{curr_row}", border=border, font=note_font, alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))

    # IMPRESIÓN PRO
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

    # IMPRESIÓN PRO
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@exportar_bp.route('/exportar/<int:pedido_id>')
def exportar_pedido(pedido_id):
    """Exporta un pedido individual con su detalle completo (Print Ready)."""
    conn = get_db()
    
    # 1. Obtener datos básicos
    row = conn.execute("SELECT * FROM pedidos WHERE id=?", (pedido_id,)).fetchone()
    if not row:
        from flask import flash, redirect, url_for
        flash("Pedido no encontrado.", "error")
        return redirect(url_for("registros.registros"))
    
    # 2. Enriquecer con el mismo helper que usa la vista detalle
    from routes.registros import _enriquecer_pedido_completo
    pedido = _enriquecer_pedido_completo(dict(row), conn)
    
    # 3. Datos de empresa para el encabezado
    empresa = conn.execute("SELECT * FROM empresa LIMIT 1").fetchone()
    if empresa: empresa = dict(empresa)
    else: empresa = {}

    # 4. Generar
    output = _crear_excel_detalle(pedido, empresa)
    
    nombre = f"Control_de_pedido_{pedido['marca'].upper().replace(' ', '_')}.xlsx"
    return send_file(output, download_name=nombre,
                     as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

